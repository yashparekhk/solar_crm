from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.core.mail import send_mail
import json
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from .models import Ticket, PublicTicket, TICKET_STATUS_CHOICES, TICKET_PRIORITY_CHOICES, TICKET_CATEGORY_CHOICES
from customers.models import Customer
from accounts.models import CustomUser

LOGIN_URL = '/accounts/login/'


@login_required(login_url=LOGIN_URL)
def tickets_list_view(request):
    search_query    = request.GET.get('q', '').strip()
    status_filter   = request.GET.get('status', '').strip()
    priority_filter = request.GET.get('priority', '').strip()

    tickets_queryset = Ticket.objects.all().order_by('-created_at')

    if search_query:
        tickets_queryset = tickets_queryset.filter(
            title__icontains=search_query
        ) | Ticket.objects.filter(
            customer__name__icontains=search_query
        )

    if status_filter:
        tickets_queryset = tickets_queryset.filter(status=status_filter)

    if priority_filter:
        tickets_queryset = tickets_queryset.filter(priority=priority_filter)

    context = {
        'tickets_list':     tickets_queryset,
        'total_count':      tickets_queryset.count(),
        'search_query':     search_query,
        'status_filter':    status_filter,
        'priority_filter':  priority_filter,
        'status_choices':   TICKET_STATUS_CHOICES,
        'priority_choices': TICKET_PRIORITY_CHOICES,
        'page_title':       'Tickets',
    }
    return render(request, 'tickets/tickets_list.html', context)


@login_required(login_url=LOGIN_URL)
def tickets_detail_view(request, ticket_id):
    ticket_instance = get_object_or_404(Ticket, pk=ticket_id)
    context = {
        'ticket_instance': ticket_instance,
        'page_title':      f'Ticket #{ticket_instance.pk}',
    }
    return render(request, 'tickets/tickets_detail.html', context)


@login_required(login_url=LOGIN_URL)
def tickets_add_view(request):
    customers_queryset = Customer.objects.all()
    users_queryset     = CustomUser.objects.all()

    if request.method == 'POST':
        new_ticket = Ticket.objects.create(
            title          = request.POST.get('title', '').strip(),
            customer_id    = request.POST.get('customer'),
            category       = request.POST.get('category', 'general'),
            priority       = request.POST.get('priority', 'medium'),
            status         = request.POST.get('status', 'open'),
            description    = request.POST.get('description', '').strip(),
            assigned_to_id = request.POST.get('assigned_to') or None,
            notes          = request.POST.get('notes', '').strip(),
        )
        messages.success(request, f'Ticket #{new_ticket.pk} created successfully!')
        return redirect('tickets_list_view')

    context = {
        'customers_list':   customers_queryset,
        'users_list':       users_queryset,
        'status_choices':   TICKET_STATUS_CHOICES,
        'priority_choices': TICKET_PRIORITY_CHOICES,
        'category_choices': TICKET_CATEGORY_CHOICES,
        'page_title':       'Add Ticket',
    }
    return render(request, 'tickets/tickets_add.html', context)


@login_required(login_url=LOGIN_URL)
def tickets_edit_view(request, ticket_id):
    ticket_instance    = get_object_or_404(Ticket, pk=ticket_id)
    customers_queryset = Customer.objects.all()
    users_queryset     = CustomUser.objects.all()

    if request.method == 'POST':
        ticket_instance.title          = request.POST.get('title', '').strip()
        ticket_instance.customer_id    = request.POST.get('customer')
        ticket_instance.category       = request.POST.get('category', 'general')
        ticket_instance.priority       = request.POST.get('priority', 'medium')
        ticket_instance.status         = request.POST.get('status', 'open')
        ticket_instance.description    = request.POST.get('description', '').strip()
        ticket_instance.assigned_to_id = request.POST.get('assigned_to') or None
        ticket_instance.notes          = request.POST.get('notes', '').strip()
        ticket_instance.save()
        messages.success(request, f'Ticket #{ticket_instance.pk} updated successfully!')
        return redirect('tickets_list_view')

    context = {
        'ticket_instance':  ticket_instance,
        'customers_list':   customers_queryset,
        'users_list':       users_queryset,
        'status_choices':   TICKET_STATUS_CHOICES,
        'priority_choices': TICKET_PRIORITY_CHOICES,
        'category_choices': TICKET_CATEGORY_CHOICES,
        'page_title':       f'Edit Ticket #{ticket_instance.pk}',
    }
    return render(request, 'tickets/tickets_edit.html', context)


@login_required(login_url=LOGIN_URL)
def tickets_delete_view(request, ticket_id):
    ticket_instance = get_object_or_404(Ticket, pk=ticket_id)
    if request.method == 'POST':
        ticket_instance.delete()
        messages.success(request, 'Ticket deleted successfully!')
        return redirect('tickets_list_view')
    context = {
        'ticket_instance': ticket_instance,
        'page_title':      'Delete Ticket',
    }
    return render(request, 'tickets/tickets_delete.html', context)


# ── PUBLIC SUPPORT PAGE ───────────────────────────────────────────────────────
def public_support_page(request):
    return render(request, 'tickets/public_support.html')


# ── PUBLIC TICKET SUBMIT ──────────────────────────────────────────────────────
@csrf_exempt
@require_POST
def public_ticket_submit(request):
    try:
        data = json.loads(request.body)

        full_name = data.get('full_name', '').strip()
        email     = data.get('email', '').strip()
        phone     = data.get('phone', '').strip()
        city      = data.get('city', '').strip() or 'Unknown'
        category  = data.get('category', 'general').strip()
        subject   = data.get('subject', '').strip()
        message   = data.get('message', '').strip()

        if not all([full_name, email, phone, subject, message]):
            return JsonResponse({'success': False, 'error': 'All fields are required.'})

        # ── Save to PublicTicket log ──
        PublicTicket.objects.create(
            full_name = full_name,
            email     = email,
            phone     = phone,
            category  = category,
            subject   = subject,
            message   = message,
        )

        # ── Get or create Customer by phone number ──
        customer, created = Customer.objects.get_or_create(
            phone = phone,
            defaults = {
                'name':        full_name,
                'email':       email,
                'address':     'Via Support Form',
                'city':        city,
                'system_size': 0.00,
            }
        )
        if not created and not customer.email:
            customer.email = email
            customer.save(update_fields=['email'])

        # ── Create real Ticket — shows in CRM tickets list ──
        ticket = Ticket.objects.create(
            title       = subject,
            customer    = customer,
            category    = category,
            priority    = 'medium',
            status      = 'open',
            description = message,
            notes       = (
                f'Submitted via public support form.\n'
                f'Name: {full_name}\n'
                f'Phone: {phone}\n'
                f'Email: {email}\n'
                f'City: {city}'
            ),
        )

        # ── Email to team ──
        send_mail(
            subject        = f'[New Support Ticket #{ticket.pk}] {subject}',
            message        = f'''A new public support ticket has been submitted.

Ticket ID : #{ticket.pk}
Name      : {full_name}
Email     : {email}
Phone     : {phone}
City      : {city}
Category  : {ticket.get_category_display()}
Subject   : {subject}

Message:
{message}
''',
            from_email     = None,
            recipient_list = ['yashparekh.k@gmail.com'],
            fail_silently  = True,
        )

        # ── Confirmation email to customer ──
        send_mail(
            subject        = f'We received your request — Ticket #{ticket.pk}',
            message        = f'''Dear {full_name},

Thank you for reaching out to Solar CRM Support.

We have received your query and our team will get back to you shortly.

Ticket Details:
───────────────
Ticket ID : #{ticket.pk}
Subject   : {subject}
Category  : {ticket.get_category_display()}

We typically respond within 24 business hours.

Regards,
Solar CRM Support Team
''',
            from_email     = None,
            recipient_list = [email],
            fail_silently  = True,
        )

        return JsonResponse({'success': True, 'ticket_id': ticket.pk})

    except Exception as e:
        return JsonResponse({'success': False, 'error': str(e)})


# ── EXPORT LEADS TO EXCEL ──────────────────────────────────────────────────────
@login_required(login_url=LOGIN_URL)
def export_leads_excel(request):
    from leads.models import Lead

    workbook  = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Leads'

    header_font  = Font(bold=True, color='FFFFFF')
    header_fill  = PatternFill(start_color='F59E0B', end_color='F59E0B', fill_type='solid')
    header_align = Alignment(horizontal='center')

    headers = ['#', 'Name', 'Phone', 'Email', 'Source', 'Status', 'Address', 'Notes', 'Created At']
    for col_num, header in enumerate(headers, 1):
        cell           = worksheet.cell(row=1, column=col_num, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = header_align

    leads_data = Lead.objects.all().order_by('-created_at')
    for row_num, lead in enumerate(leads_data, 2):
        worksheet.cell(row=row_num, column=1, value=row_num - 1)
        worksheet.cell(row=row_num, column=2, value=lead.name)
        worksheet.cell(row=row_num, column=3, value=lead.phone)
        worksheet.cell(row=row_num, column=4, value=lead.email)
        worksheet.cell(row=row_num, column=5, value=lead.source)
        worksheet.cell(row=row_num, column=6, value=lead.get_status_display())
        worksheet.cell(row=row_num, column=7, value=lead.address)
        worksheet.cell(row=row_num, column=8, value=lead.notes)
        worksheet.cell(row=row_num, column=9, value=str(lead.created_at.strftime('%d %b %Y')))

    for col in worksheet.columns:
        max_len = max(len(str(cell.value or '')) for cell in col)
        worksheet.column_dimensions[col[0].column_letter].width = max_len + 4

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="solar_crm_leads.xlsx"'
    workbook.save(response)
    return response


@login_required(login_url=LOGIN_URL)
def import_leads_view(request):
    if request.method == 'POST' and request.FILES.get('excel_file'):
        excel_file     = request.FILES['excel_file']
        imported_count = 0
        error_rows     = []

        try:
            workbook  = openpyxl.load_workbook(excel_file)
            worksheet = workbook.active

            header_row = [str(cell.value or '').strip().lower() for cell in worksheet[1]]

            def get_col(names):
                for name in names:
                    for i, h in enumerate(header_row):
                        if name in h:
                            return i
                return None

            col_name    = get_col(['name', 'full name'])
            col_phone   = get_col(['phone', 'mobile', 'contact'])
            col_email   = get_col(['email', 'mail'])
            col_source  = get_col(['source'])
            col_status  = get_col(['status'])
            col_address = get_col(['address'])
            col_notes   = get_col(['notes', 'note'])

            if col_name is None:
                col_name = 0; col_phone = 1; col_email = 2
                col_source = 3; col_status = 4; col_address = 5; col_notes = 6

            for row_num, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), 2):
                try:
                    def get_val(col):
                        if col is not None and col < len(row) and row[col]:
                            val = str(row[col]).strip()
                            return '' if val.lower() == 'none' else val
                        return ''

                    lead_name    = get_val(col_name)
                    lead_phone   = get_val(col_phone)
                    lead_email   = get_val(col_email)
                    lead_source  = get_val(col_source)
                    lead_status  = get_val(col_status).lower()
                    lead_address = get_val(col_address)
                    lead_notes   = get_val(col_notes)

                    if not lead_name and not lead_phone:
                        continue
                    if not lead_name or not lead_phone:
                        error_rows.append(f'Row {row_num}: Name and Phone are required')
                        continue

                    valid_statuses = ['new', 'contacted', 'qualified', 'lost']
                    if lead_status not in valid_statuses:
                        lead_status = 'new'

                    from leads.models import Lead
                    Lead.objects.create(
                        name=lead_name, phone=lead_phone, email=lead_email,
                        source=lead_source, status=lead_status,
                        address=lead_address, notes=lead_notes,
                    )
                    imported_count += 1

                except Exception as row_error:
                    error_rows.append(f'Row {row_num}: {str(row_error)}')

            if imported_count > 0:
                messages.success(request, f'Successfully imported {imported_count} lead(s)!')
            if error_rows:
                for error in error_rows[:5]:
                    messages.warning(request, error)
            if imported_count == 0 and not error_rows:
                messages.warning(request, 'No data found in the file.')

        except Exception as file_error:
            messages.error(request, f'Error reading file: {str(file_error)}')

        return redirect('leads_list_view')

    return render(request, 'leads/leads_import.html', {'page_title': 'Import Leads'})


# ── DOWNLOAD IMPORT TEMPLATE ──────────────────────────────────────────────────
@login_required(login_url=LOGIN_URL)
def download_import_template(request):
    workbook  = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = 'Leads Import Template'

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='F59E0B', end_color='F59E0B', fill_type='solid')

    headers = ['Name *', 'Phone *', 'Email', 'Source', 'Status', 'Address', 'Notes']
    for col_num, header in enumerate(headers, 1):
        cell      = worksheet.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        worksheet.column_dimensions[cell.column_letter].width = 20

    sample_row = ['Yash Parekh', '9876543210', 'yash@gmail.com', 'Website', 'new', 'Surat, Gujarat', 'Interested in 5kW system']
    for col_num, value in enumerate(sample_row, 1):
        worksheet.cell(row=2, column=col_num, value=value)

    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename="leads_import_template.xlsx"'
    workbook.save(response)
    return response